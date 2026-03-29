import mysql.connector
import pandas as pd
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Database configuration
db_config = {
    'host': 'localhost',
    'user': 'root',
    'password': '',
    'database': 'attendance_system'
}

def connect_to_database():
    """Establish database connection."""
    try:
        conn = mysql.connector.connect(**db_config)
        print("Database connection successful.")
        return conn
    except mysql.connector.Error as err:
        print(f"Database Error: {err}")
        return None

def get_attendance_data(conn, start_date=None, end_date=None):
    """Fetch attendance data from database."""
    cursor = conn.cursor()
    
    # Base query
    query = """
    SELECT 
        s.roll_number,
        s.name,
        DATE(a.date_time) as date,
        TIME(a.date_time) as time
    FROM 
        students s
        LEFT JOIN attendance a ON s.roll_number = a.roll_number
    """
    
    # Add date filters if provided
    if start_date and end_date:
        query += " WHERE DATE(a.date_time) BETWEEN %s AND %s"
        cursor.execute(query, (start_date, end_date))
    else:
        cursor.execute(query)
    
    # Fetch all records
    records = cursor.fetchall()
    cursor.close()
    
    return records

def create_excel_report(data, output_file):
    """Create formatted Excel report from attendance data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4A69BD", end_color="4A69BD", fill_type="solid")
    
    # Write headers
    headers = ["Roll Number", "Name", "Date", "Time", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Write data
    for row, record in enumerate(data, 2):
        roll_number, name, date, time = record
        
        ws.cell(row=row, column=1, value=roll_number)
        ws.cell(row=row, column=2, value=name)
        
        # Handle date formatting safely
        if date:
            if hasattr(date, 'strftime'):
                date_str = date.strftime('%Y-%m-%d')
            else:
                date_str = str(date)
        else:
            date_str = ""
        ws.cell(row=row, column=3, value=date_str)
        
        # Handle time formatting safely
        if time:
            if hasattr(time, 'strftime'):
                time_str = time.strftime('%H:%M:%S')
            else:
                time_str = str(time)
        else:
            time_str = ""
        ws.cell(row=row, column=4, value=time_str)
        
        ws.cell(row=row, column=5, value="Present" if time else "Absent")
        
        # Center align all cells in the row
        for col in range(1, 6):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
    
    # Auto-adjust column widths
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Create 'exports' directory if it doesn't exist
    if not os.path.exists('exports'):
        os.makedirs('exports')
    
    # Save the workbook
    wb.save(output_file)
    print(f"Report saved successfully to {output_file}")

def main():
    # Connect to database
    conn = connect_to_database()
    if not conn:
        print("Failed to connect to database. Exiting.")
        return
    
    try:
        # Get user input for date range
        print("\nAttendance Export Tool")
        print("---------------------")
        print("1. Export all attendance data")
        print("2. Export data for specific date range")
        choice = input("Enter your choice (1/2): ")
        
        if choice == "2":
            start_date = input("Enter start date (YYYY-MM-DD): ")
            end_date = input("Enter end date (YYYY-MM-DD): ")
            data = get_attendance_data(conn, start_date, end_date)
            date_range = f"{start_date}_to_{end_date}"
        else:
            data = get_attendance_data(conn)
            date_range = "all_time"
        
        # Generate output filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f"exports/attendance_report_{date_range}_{timestamp}.xlsx"
        
        # Create and save the report
        create_excel_report(data, output_file)
        
    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        conn.close()
        print("\nDatabase connection closed.")

if __name__ == "__main__":
    main()