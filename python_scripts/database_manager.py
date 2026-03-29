import mysql.connector
import pandas as pd
from datetime import datetime, timedelta
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Database configuration
db_config = {
    'host': 'localhost',
    'user': 'root',
    'password': '',
    'database': 'attendance_system'
}

# Excel file configuration
EXCEL_FILE = "attendance_records.xlsx"

def connect_to_database():
    """Establish database connection."""
    try:
        conn = mysql.connector.connect(**db_config)
        print("Database connection successful.")
        return conn
    except mysql.connector.Error as err:
        print(f"Database Error: {err}")
        return None

def create_excel_file():
    """Create Excel file with proper headers if it doesn't exist."""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Records"
        
        # Define headers
        headers = ["Roll Number", "Name", "Date", "Time", "Status"]
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4A69BD", end_color="4A69BD", fill_type="solid")
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        wb.save(EXCEL_FILE)
        print(f"Created new Excel file: {EXCEL_FILE}")

def add_student_to_db(conn, roll_number, name):
    """Add a new student to the database."""
    cursor = conn.cursor()
    try:
        # Check if student already exists
        cursor.execute("SELECT roll_number FROM students WHERE roll_number = %s", (roll_number,))
        if cursor.fetchone():
            print(f"Student with Roll Number {roll_number} already exists.")
            return False
        
        # Insert new student
        sql = "INSERT INTO students (roll_number, name) VALUES (%s, %s)"
        val = (roll_number, name)
        cursor.execute(sql, val)
        conn.commit()
        print(f"Student {name} ({roll_number}) added to database successfully.")
        return True
    except mysql.connector.Error as err:
        print(f"Error adding student: {err}")
        conn.rollback()
        return False
    finally:
        cursor.close()

def remove_student_from_db(conn, roll_number):
    """Remove a student from the database."""
    cursor = conn.cursor()
    try:
        # Check if student exists
        cursor.execute("SELECT name FROM students WHERE roll_number = %s", (roll_number,))
        result = cursor.fetchone()
        if not result:
            print(f"Student with Roll Number {roll_number} not found.")
            return False
        
        name = result[0]
        
        # Delete attendance records first (due to foreign key constraint)
        cursor.execute("DELETE FROM attendance WHERE roll_number = %s", (roll_number,))
        
        # Delete student
        cursor.execute("DELETE FROM students WHERE roll_number = %s", (roll_number,))
        conn.commit()
        print(f"Student {name} ({roll_number}) removed from database successfully.")
        return True
    except mysql.connector.Error as err:
        print(f"Error removing student: {err}")
        conn.rollback()
        return False
    finally:
        cursor.close()

def add_attendance_record(conn, roll_number, date_time=None):
    """Add an attendance record to the database."""
    cursor = conn.cursor()
    try:
        # Check if student exists
        cursor.execute("SELECT name FROM students WHERE roll_number = %s", (roll_number,))
        result = cursor.fetchone()
        if not result:
            print(f"Student with Roll Number {roll_number} not found.")
            return False
        
        name = result[0]
        
        # Use current time if not provided
        if date_time is None:
            date_time = datetime.now()
        
        # Check if attendance already exists for this date
        date_str = date_time.strftime('%Y-%m-%d')
        cursor.execute("SELECT id FROM attendance WHERE roll_number = %s AND DATE(date_time) = %s", 
                      (roll_number, date_str))
        if cursor.fetchone():
            print(f"Attendance for {roll_number} already exists on {date_str}")
            return False
        
        # Insert attendance record
        sql = "INSERT INTO attendance (roll_number, date_time) VALUES (%s, %s)"
        val = (roll_number, date_time.strftime('%Y-%m-%d %H:%M:%S'))
        cursor.execute(sql, val)
        conn.commit()
        print(f"Attendance marked for {name} ({roll_number}) at {date_time.strftime('%Y-%m-%d %H:%M:%S')}")
        return True
    except mysql.connector.Error as err:
        print(f"Error adding attendance: {err}")
        conn.rollback()
        return False
    finally:
        cursor.close()

def remove_attendance_record(conn, roll_number, date):
    """Remove an attendance record from the database."""
    cursor = conn.cursor()
    try:
        # Check if attendance record exists
        cursor.execute("SELECT s.name FROM attendance a JOIN students s ON a.roll_number = s.roll_number WHERE a.roll_number = %s AND DATE(a.date_time) = %s", 
                      (roll_number, date))
        result = cursor.fetchone()
        if not result:
            print(f"No attendance record found for {roll_number} on {date}")
            return False
        
        name = result[0]
        
        # Delete attendance record
        cursor.execute("DELETE FROM attendance WHERE roll_number = %s AND DATE(date_time) = %s", 
                      (roll_number, date))
        conn.commit()
        print(f"Attendance record removed for {name} ({roll_number}) on {date}")
        return True
    except mysql.connector.Error as err:
        print(f"Error removing attendance: {err}")
        conn.rollback()
        return False
    finally:
        cursor.close()

def view_all_students(conn):
    """Display all students in the database."""
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT roll_number, name FROM students ORDER BY roll_number")
        students = cursor.fetchall()
        
        if not students:
            print("No students found in database.")
            return
        
        print("\n" + "="*50)
        print("STUDENTS IN DATABASE")
        print("="*50)
        print(f"{'Roll Number':<15} {'Name':<30}")
        print("-" * 50)
        for roll_number, name in students:
            print(f"{roll_number:<15} {name:<30}")
        print("="*50)
        print(f"Total Students: {len(students)}")
        
    except mysql.connector.Error as err:
        print(f"Error fetching students: {err}")
    finally:
        cursor.close()

def view_attendance_records(conn, date=None):
    """Display attendance records for a specific date or all records."""
    cursor = conn.cursor()
    try:
        if date:
            sql = """
            SELECT s.roll_number, s.name, DATE(a.date_time), TIME(a.date_time)
            FROM students s
            LEFT JOIN attendance a ON s.roll_number = a.roll_number AND DATE(a.date_time) = %s
            ORDER BY s.roll_number
            """
            cursor.execute(sql, (date,))
        else:
            sql = """
            SELECT s.roll_number, s.name, DATE(a.date_time), TIME(a.date_time)
            FROM students s
            LEFT JOIN attendance a ON s.roll_number = a.roll_number
            ORDER BY s.roll_number, a.date_time
            """
            cursor.execute(sql)
        
        records = cursor.fetchall()
        
        if not records:
            print("No attendance records found.")
            return
        
        print("\n" + "="*70)
        print("ATTENDANCE RECORDS")
        print("="*70)
        print(f"{'Roll Number':<15} {'Name':<20} {'Date':<12} {'Time':<10} {'Status':<10}")
        print("-" * 70)
        
        for roll_number, name, date, time in records:
            status = "Present" if time else "Absent"
            date_str = date.strftime('%Y-%m-%d') if date else "-"
            
            # Handle time object (could be timedelta or datetime.time)
            if time:
                if hasattr(time, 'strftime'):  # datetime.time object
                    time_str = time.strftime('%H:%M:%S')
                else:  # timedelta object
                    total_seconds = int(time.total_seconds())
                    hours = total_seconds // 3600
                    minutes = (total_seconds % 3600) // 60
                    seconds = total_seconds % 60
                    time_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
            else:
                time_str = "-"
                
            print(f"{roll_number:<15} {name:<20} {date_str:<12} {time_str:<10} {status:<10}")
        
        print("="*70)
        
    except mysql.connector.Error as err:
        print(f"Error fetching attendance: {err}")
    finally:
        cursor.close()

def sync_database_to_excel(conn):
    """Sync all database records to Excel file."""
    cursor = conn.cursor()
    try:
        # Get all attendance data
        sql = """
        SELECT s.roll_number, s.name, DATE(a.date_time), TIME(a.date_time)
        FROM students s
        LEFT JOIN attendance a ON s.roll_number = a.roll_number
        ORDER BY s.roll_number, a.date_time
        """
        cursor.execute(sql)
        records = cursor.fetchall()
        
        # Create or load Excel file
        create_excel_file()
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # Clear existing data (keep headers)
        for row in range(2, ws.max_row + 1):
            for col in range(1, 6):
                ws.cell(row=row, column=col).value = None
        
        # Write data
        for row, record in enumerate(records, 2):
            roll_number, name, date, time = record
            
            ws.cell(row=row, column=1, value=roll_number)
            ws.cell(row=row, column=2, value=name)
            ws.cell(row=row, column=3, value=date.strftime('%Y-%m-%d') if date else "")
            
            # Handle time object (could be timedelta or datetime.time)
            if time:
                if hasattr(time, 'strftime'):  # datetime.time object
                    time_str = time.strftime('%H:%M:%S')
                else:  # timedelta object
                    total_seconds = int(time.total_seconds())
                    hours = total_seconds // 3600
                    minutes = (total_seconds % 3600) // 60
                    seconds = total_seconds % 60
                    time_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
            else:
                time_str = ""
                
            ws.cell(row=row, column=4, value=time_str)
            ws.cell(row=row, column=5, value="Present" if time else "Absent")
            
            # Center align all cells in the row
            for col in range(1, 6):
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        
        wb.save(EXCEL_FILE)
        print(f"Database synchronized to Excel file: {EXCEL_FILE}")
        print(f"Total records synced: {len(records)}")
        
    except mysql.connector.Error as err:
        print(f"Error syncing to Excel: {err}")
    finally:
        cursor.close()

def add_record_to_excel(roll_number, name, date, time, status):
    """Add a single record to Excel file."""
    try:
        # Create file if it doesn't exist
        create_excel_file()
        
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # Find next empty row
        next_row = ws.max_row + 1
        
        # Add record
        ws.cell(row=next_row, column=1, value=roll_number)
        ws.cell(row=next_row, column=2, value=name)
        ws.cell(row=next_row, column=3, value=date)
        ws.cell(row=next_row, column=4, value=time)
        ws.cell(row=next_row, column=5, value=status)
        
        # Center align
        for col in range(1, 6):
            ws.cell(row=next_row, column=col).alignment = Alignment(horizontal='center')
        
        wb.save(EXCEL_FILE)
        print(f"Record added to Excel: {roll_number} - {name} - {date} - {time} - {status}")
        
    except Exception as e:
        print(f"Error adding record to Excel: {e}")

def main():
    """Main menu for database management."""
    conn = connect_to_database()
    if not conn:
        print("Failed to connect to database. Exiting.")
        return
    
    while True:
        print("\n" + "="*60)
        print("SMART ATTENDANCE SYSTEM - DATABASE MANAGER")
        print("="*60)
        print("1. Add Student")
        print("2. Remove Student")
        print("3. Add Attendance Record")
        print("4. Remove Attendance Record")
        print("5. View All Students")
        print("6. View Attendance Records")
        print("7. Sync Database to Excel")
        print("8. Exit")
        print("="*60)
        
        choice = input("Enter your choice (1-8): ").strip()
        
        if choice == "1":
            roll_number = input("Enter Roll Number: ").strip()
            name = input("Enter Student Name: ").strip()
            if roll_number and name:
                add_student_to_db(conn, roll_number, name)
            else:
                print("Roll Number and Name cannot be empty.")
        
        elif choice == "2":
            roll_number = input("Enter Roll Number to remove: ").strip()
            if roll_number:
                confirm = input(f"Are you sure you want to remove student {roll_number}? (y/n): ").lower()
                if confirm == 'y':
                    remove_student_from_db(conn, roll_number)
            else:
                print("Roll Number cannot be empty.")
        
        elif choice == "3":
            roll_number = input("Enter Roll Number: ").strip()
            if roll_number:
                date_input = input("Enter date (YYYY-MM-DD) or press Enter for today: ").strip()
                time_input = input("Enter time (HH:MM:SS) or press Enter for now: ").strip()
                
                try:
                    if date_input:
                        date_time = datetime.strptime(f"{date_input} {time_input or '00:00:00'}", "%Y-%m-%d %H:%M:%S")
                    else:
                        date_time = datetime.now()
                    
                    if add_attendance_record(conn, roll_number, date_time):
                        # Also add to Excel
                        cursor = conn.cursor()
                        cursor.execute("SELECT name FROM students WHERE roll_number = %s", (roll_number,))
                        result = cursor.fetchone()
                        if result:
                            name = result[0]
                            add_record_to_excel(roll_number, name, 
                                              date_time.strftime('%Y-%m-%d'),
                                              date_time.strftime('%H:%M:%S'),
                                              "Present")
                        cursor.close()
                except ValueError:
                    print("Invalid date/time format. Use YYYY-MM-DD HH:MM:SS")
            else:
                print("Roll Number cannot be empty.")
        
        elif choice == "4":
            roll_number = input("Enter Roll Number: ").strip()
            date = input("Enter date (YYYY-MM-DD): ").strip()
            if roll_number and date:
                remove_attendance_record(conn, roll_number, date)
            else:
                print("Roll Number and Date cannot be empty.")
        
        elif choice == "5":
            view_all_students(conn)
        
        elif choice == "6":
            date = input("Enter date to view (YYYY-MM-DD) or press Enter for all: ").strip()
            if date:
                view_attendance_records(conn, date)
            else:
                view_attendance_records(conn)
        
        elif choice == "7":
            sync_database_to_excel(conn)
        
        elif choice == "8":
            print("Exiting Database Manager...")
            break
        
        else:
            print("Invalid choice. Please enter a number between 1-8.")
    
    conn.close()
    print("Database connection closed.")

if __name__ == "__main__":
    main()